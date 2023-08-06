import openpyxl
import requests
import threading
import sys
from tqdm import tqdm
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


# 获取xlsx中的数据，存到列表中
def open_xlsx():
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(r"ipport.xlsx")
    except Exception as e:
        print("error:请创建一个ipport.xlsx的文件，将数据存放到里面")
        sys.exit(1)
    # 选择第一个工作表
    sheet = workbook.worksheets[0]

    max_row = sheet.max_row

    data = []

    for row in range(1, max_row + 1):
        column1_value = sheet.cell(row=row, column=1).value
        column2_value = sheet.cell(row=row, column=2).value

        if column1_value is not None and column2_value is not None:
            combined_value = f'{column1_value}://{column2_value}'
            data.append(combined_value)

    return data


# 去请求网址
def request_ipport(ipport,waittime,result):
    try:
        # 发送 GET 请求
        response = requests.get(ipport,timeout=waittime)
        # 获取响应状态码
        status_code = response.status_code
        # 使用BeautifulSoup解析网页内容
        soup = BeautifulSoup(response.content, 'html.parser')
        # 提取网页标题
        title = soup.title.string
        result.append([ipport,title])

    except Exception as e:
        pass

# 保存数据到result.xlsx中
def save_xlsx(data):
    # 检测文件是否存在
    try:
        workbook = load_workbook('result.xlsx')
        # 清除现有的内容
        workbook.active.delete_rows(1, workbook.active.max_row)
    except FileNotFoundError:
        # 文件不存在，创建新的Workbook对象
        workbook = Workbook()
    # 选择或创建一个新的Sheet
    sheet = workbook.active
    # 写入数据
    for item in data:
        sheet.append(item)
    # 保存文件
    workbook.save('result.xlsx')

# 多线程处理
def process_data_in_threads(data_list,MAX_THREADS,waittime):
    result = []
    thread_pool = []

    for data in data_list:
        thread = threading.Thread(target=request_ipport, args=(data,waittime,result))
        thread_pool.append(thread)
        thread.start()

        while len(thread_pool) >= MAX_THREADS:
            for t in thread_pool:
                if not t.is_alive():
                    thread_pool.remove(t)
                    break

    for thread in thread_pool:
        thread.join()

    return result

